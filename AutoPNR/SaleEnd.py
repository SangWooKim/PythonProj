#-*- coding: utf-8 -*-

import configparser
import logging.handlers
import os
import AutoPNRLib
import openpyxl

#로그 정의
logdir = '{0}\\log'.format(os.getcwd())

if not os.path.exists(logdir):
    os.makedirs(logdir)

logger = logging.getLogger('auto_pnr_log')
fileHandler = logging.FileHandler('{0}/autoPNR.log'.format(logdir))
streamHandler = logging.StreamHandler()

logger.addHandler(fileHandler)
logger.addHandler(streamHandler)
logger.setLevel(logging.DEBUG)
'''
로직 순서
1. 주어진 엑셀 파일에서 F 컬럼 7번째 로우 부터 모두 읽는다.
2. 읽은 데이타에서 "날짜" 단어로 구분한다.
  * XXXXXXXX구간:제주→김포,날짜:2/1 (목),시간:09:05-1개 -->날짜:2/1 (목),시간:09:05-1개
3. 날짜를 기준으로 정렬한다.
4. 정렬된 row를 F 컬럼부터 하나씩 읽어서 result 시트에 저장한다.
'''
if __name__ == "__main__":
    logger.info('=======================================')
    logger.info('Sale END excel')
    logger.info('=======================================')

    config = configparser.ConfigParser()
    if os.path.exists('.\config.txt'):
        config.read('.\config.txt')

    # 설정 파일 확인
    file_path = 'C:\\Users\\swkim\\Documents\\카카오톡 받은 파일\\SomSaleEndListExcel_test.xlsx'
    start_row = 7
    column = 2
    sheet_name = 'Sheet'



    #엑셀 파일에서 PNR 코드만 읽어서 리스트를 만들어 준다.
    basicfunction = AutoPNRLib.AutoPNRLib()
    option_list = basicfunction.LoadXLSXData(file_path, sheet_name, column, start_row)

    option_list.sort()
    #print(option_list)


    #리스트 확인
    #print(len(pnr_code_list))
    #print(option_list)


    wb = openpyxl.load_workbook(file_path)
    ws_result = wb.create_sheet('result')
    ws_data = wb[sheet_name]

    col_range = 'CDEFGHIJKLMNOPQRSTUVW'


    # A ~ W 까지
    for idx, option in enumerate(option_list) :
        row_index = option[1]
        for idx_col, arg in enumerate(col_range):
            if idx_col == 0:
                temp_list = ws_data[arg + str(row_index)].value.split('구간')[1].split('시간')
                ws_result.cell(row=idx + start_row, column=1, value='구간' + temp_list[0])
            ws_result.cell(row=idx + start_row, column=idx_col + 2, value=ws_data[arg + str(row_index)].value)

    wb.save(file_path)
    wb.close()